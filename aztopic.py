import nltk
nltk.download("stopwords")

import numpy as np
from openpyxl import load_workbook
import json
from collections import OrderedDict
from itertools import islice
import glob

import spacy
from spacy.tokens import DocBin
from nltk.corpus import stopwords

import warnings
warnings.filterwarnings("ignore", category=DepreciationWarning)

stopwords = stopwords.words('english')

def load_data(excelSheet):
    wb = load_workbook(excelSheet)
    sh = wb['CS Feedback']

    data_list = []

    for row in islice(sh.values, 1, sh.max_row):
        feedback = OrderedDict()
        feedback['date'] = row[0]
        feedback['issue'] = row[1]
        feedback['product'] = row[2]
        feedback['name'] = row[3]
        feedback['email'] = row[4]
        feedback['comment'] = row[5]
        feedback['ip'] = row[6]
        feedback['jsession'] = row[7]
        feedback['followup'] = row[8]

    j = json.dumps(data_list)

    return j

def make_docs(data):
    docs = []
    for doc,label in nip.pip(data, as_tuples = True):
        

